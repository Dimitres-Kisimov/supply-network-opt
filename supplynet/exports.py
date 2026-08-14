"""Executive deliverables: a PDF report and an Excel workbook.

The PDF (matplotlib PdfPages) is a "network atlas": eight numbered plates that
share one chrome (see :mod:`supplynet.plates` / :mod:`supplynet.atlas`) -- a
cover with the synthetic-data disclaimer and headline savings, a network map of
opened DCs and their flows, a cost-breakdown bar, a safety-stock pooling chart,
the CO2 Pareto sweep, the N-1 resilience screen, the service-level frontier and
the demand-growth expansion staircase. The Excel workbook (openpyxl) has
Summary, Facilities, Flows and SafetyStock sheets.
"""

import os
import textwrap

import matplotlib

matplotlib.use("Agg")
# Dollar signs in our labels are literal currency, not math delimiters.
matplotlib.rcParams["text.parse_math"] = False
import matplotlib.pyplot as plt  # noqa: E402
from matplotlib.backends.backend_pdf import PdfPages  # noqa: E402
from matplotlib.lines import Line2D  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

from supplynet import plates  # noqa: E402
from supplynet.atlas import (  # noqa: E402
    AQUA,
    BLUE,
    BLUE_DARK,
    BLUE_FLOW,
    BLUE_SOFT,
    CRITICAL,
    GOOD_TEXT,
    GRAY_FILL,
    INK,
    INK2,
    MUTED,
    ORANGE,
    PAPER,
    RULE,
    SERIES_GRAY,
)
from supplynet.co2_sensitivity import to_csv, to_svg, tradeoff_readout  # noqa: E402
from supplynet.growth import growth_readout  # noqa: E402
from supplynet.growth import to_csv as growth_to_csv  # noqa: E402
from supplynet.growth import to_svg as growth_to_svg  # noqa: E402
from supplynet.pipeline import PipelineResult, run_pipeline  # noqa: E402
from supplynet.resilience import resilience_readout  # noqa: E402
from supplynet.service_frontier import frontier_readout  # noqa: E402
from supplynet.service_frontier import to_csv as frontier_to_csv  # noqa: E402
from supplynet.service_frontier import to_svg as frontier_to_svg  # noqa: E402

DISCLAIMER = (
    "All data in this report is SYNTHETIC and generated from a fixed random seed. "
    "Figures are model-based estimates from a facility-location MILP, a min-cost "
    "network-flow model and textbook safety-stock formulas under stated "
    "assumptions (normal, independent demand; fixed lead times). They do not "
    "describe any real company and are not guarantees."
)


def _cover_page(pdf: PdfPages, r: PipelineResult) -> None:
    fig = plates.new_plate(1, title=None)

    fig.text(0.5, 0.855, "Supply-Network Optimization", ha="center",
             fontsize=24, fontweight="bold", color=INK)
    fig.text(0.5, 0.815, "Distribution-Center Siting, Network Flow & Safety Stock",
             ha="center", fontsize=12.5, color=INK2)

    headline = (
        f"Optimized network opens {r.milp.n_opened} of {r.data.n_dcs} candidate DCs "
        f"at a total fixed+transport cost of ${r.milp.total_cost:,.0f}."
    )
    saved = (
        f"That is ${r.savings_abs:,.0f} ({r.savings_pct:.1f}%) below the "
        f"greedy cheapest-open baseline (${r.greedy.total_cost:,.0f})."
    )
    pool = (
        f"Risk pooling cuts modeled safety stock by "
        f"{r.pooling.network_reduction_pct:.0f}% "
        f"(network) / {r.pooling.reduction_pct:.0f}% (full centralization) "
        f"at a {r.service_level:.0%} service level."
    )
    fig.text(0.5, 0.745, textwrap.fill(headline, 90), ha="center", fontsize=12,
             color=INK)
    fig.text(0.5, 0.705, textwrap.fill(saved, 90), ha="center", fontsize=12,
             color=GOOD_TEXT)
    fig.text(0.5, 0.665, textwrap.fill(pool, 90), ha="center", va="top",
             fontsize=12, color=INK2)

    # Headline metric tiles: white cards behind a hairline ring.
    tiles = [
        (f"{r.milp.n_opened}", "DCs opened"),
        (f"{r.savings_pct:.1f}%", "cost vs baseline"),
        (f"${r.flow['lp'].total_cost:,.0f}", "min-cost flow"),
        (f"{r.pooling.network_reduction_pct:.0f}%", "safety-stock pooling"),
    ]
    for k, (val, lab) in enumerate(tiles):
        x0 = 0.075 + k * 0.2225
        fig.add_artist(plt.Rectangle((x0, 0.42), 0.20, 0.145,
                                     transform=fig.transFigure, facecolor=PAPER,
                                     edgecolor=RULE, linewidth=0.9))
        fig.text(x0 + 0.10, 0.505, val, ha="center", fontsize=19,
                 fontweight="bold", color=INK)
        fig.text(x0 + 0.10, 0.45, lab, ha="center", fontsize=8.5, color=MUTED)

    # Plate index: the atlas contents, one line per row of plates.
    fig.text(0.5, 0.345, "PLATE 02 NETWORK MAP  -  03 COST BREAKDOWN  -  "
             "04 RISK POOLING  -  05 COST VS CO2",
             ha="center", fontsize=8, color=MUTED)
    fig.text(0.5, 0.315, "PLATE 06 N-1 RESILIENCE  -  07 SERVICE FRONTIER  -  "
             "08 GROWTH STAIRCASE",
             ha="center", fontsize=8, color=MUTED)

    fig.text(0.5, 0.19, textwrap.fill(DISCLAIMER, 92), ha="center", va="center",
             fontsize=9, color=INK2,
             bbox=dict(boxstyle="round,pad=0.6", facecolor="#f9f9f7",
                       edgecolor=RULE, linewidth=0.8))
    fig.add_artist(Line2D([plates.MARGIN_L, plates.MARGIN_R], [0.09, 0.09],
                          transform=fig.transFigure, color=RULE, linewidth=0.8))
    fig.text(0.5, 0.055, "Synthetic data - seed "
             f"{r.data.seed} - generated for portfolio demonstration",
             ha="center", fontsize=8, color=MUTED)
    pdf.savefig(fig)
    plt.close(fig)


def _network_map(pdf: PdfPages, r: PipelineResult) -> None:
    fig = plates.new_plate(2, "Optimized Network: opened DCs and outbound flows")
    ax = fig.add_axes([0.07, 0.16, 0.60, 0.68])
    plates.despine(ax)
    d = r.data
    opened = set(r.milp.opened)
    assign = r.milp.assignment
    dc_ids = list(d.dcs["dc_id"])

    def dc_size(cap: float) -> float:
        # Node size by capacity (4k-9k units -> 60-300 pt^2).
        return 60.0 + 240.0 * (cap - 4000.0) / 5000.0

    # Flows: opened DC -> customer, line width by shipped volume.
    max_flow = assign.max() if assign.size else 1.0
    for i, dc_id in enumerate(dc_ids):
        if dc_id not in opened:
            continue
        for j in range(d.n_customers):
            f = assign[i, j]
            if f <= 1e-6:
                continue
            ax.plot(
                [d.dcs.iloc[i]["x"], d.customers.iloc[j]["x"]],
                [d.dcs.iloc[i]["y"], d.customers.iloc[j]["y"]],
                color=BLUE_FLOW, linewidth=0.5 + 2.8 * f / max_flow, zorder=1,
                solid_capstyle="round",
            )

    cust_sizes = 10.0 + 30.0 * (d.customers["demand_mean"] - 200.0) / 800.0
    ax.scatter(d.customers["x"], d.customers["y"], s=cust_sizes, c=SERIES_GRAY,
               marker="o", linewidths=0.6, edgecolors=PAPER, zorder=2)

    closed = d.dcs[~d.dcs["dc_id"].isin(opened)]
    open_dcs = d.dcs[d.dcs["dc_id"].isin(opened)]
    ax.scatter(closed["x"], closed["y"], s=dc_size(closed["capacity"]),
               facecolors="none", edgecolors=RULE, linewidths=1.2, marker="s",
               zorder=3)
    ax.scatter(open_dcs["x"], open_dcs["y"], s=dc_size(open_dcs["capacity"]),
               c=BLUE, marker="s", linewidths=1.4, edgecolors=PAPER, zorder=4)
    ax.scatter(d.plants["x"], d.plants["y"], s=190, c=ORANGE, marker="^",
               linewidths=1.4, edgecolors=PAPER, zorder=5)

    # Direct labels: every facility is named on the map.
    for _, row in open_dcs.iterrows():
        ax.annotate(row["dc_id"], (row["x"], row["y"]), fontsize=9,
                    fontweight="bold", color=INK,
                    xytext=(9, 8), textcoords="offset points")
    for _, row in closed.iterrows():
        ax.annotate(row["dc_id"], (row["x"], row["y"]), fontsize=7.5,
                    color=MUTED, xytext=(7, 6), textcoords="offset points")
    for _, row in d.plants.iterrows():
        ax.annotate(row["plant_id"], (row["x"], row["y"]), fontsize=8,
                    color=INK2, xytext=(6, -10), textcoords="offset points")

    ax.set_aspect("equal")
    ax.set_xlim(-4, 104)
    ax.set_ylim(-4, 104)
    ax.set_xlabel("x (km)")
    ax.set_ylabel("y (km)")
    ax.grid(True)
    ax.set_axisbelow(True)

    # Legend column to the right of the map, with fixed-size key marks.
    handles = [
        Line2D([], [], marker="o", linestyle="none", markersize=5,
               markerfacecolor=SERIES_GRAY, markeredgecolor=PAPER,
               label="Customer zone"),
        Line2D([], [], marker="s", linestyle="none", markersize=8,
               markerfacecolor="none", markeredgecolor=RULE,
               label="DC (not opened)"),
        Line2D([], [], marker="s", linestyle="none", markersize=9,
               markerfacecolor=BLUE, markeredgecolor=PAPER, label="DC (opened)"),
        Line2D([], [], marker="^", linestyle="none", markersize=9,
               markerfacecolor=ORANGE, markeredgecolor=PAPER, label="Plant"),
        Line2D([], [], color=BLUE_FLOW, linewidth=2.0, label="Outbound flow"),
    ]
    ax.legend(handles=handles, loc="upper left", bbox_to_anchor=(1.06, 1.0),
              borderaxespad=0.0)
    fig.text(0.705, 0.55, "Square size = DC capacity\n"
             "Dot size = zone demand\nLine width = shipped units",
             fontsize=8, color=MUTED, va="top", linespacing=1.7)

    plates.footer(fig, "Marker size scales with DC capacity (squares) and zone "
                  "demand (dots); flow-line width with shipped units. Synthetic, "
                  "seeded coordinates on a 100 km grid; model-based estimates.")
    pdf.savefig(fig)
    plt.close(fig)


def _cost_breakdown(pdf: PdfPages, r: PipelineResult) -> None:
    fig = plates.new_plate(
        3, "Cost breakdown",
        f"MILP saves ${r.savings_abs:,.0f} ({r.savings_pct:.1f}%) "
        "vs greedy baseline",
    )
    ax = fig.add_axes([0.26, 0.17, 0.40, 0.60])
    plates.despine(ax)
    labels = ["Greedy baseline", "MILP optimized"]
    fixed = [r.greedy.fixed_cost, r.milp.fixed_cost]
    transport = [r.greedy.transport_cost, r.milp.transport_cost]

    # Emphasis is ink-vs-gray: the optimized bar wears the blues, the baseline
    # wears grays. Stack order is identical (fixed below, transport above) and
    # a surface gap separates the segments.
    ax.bar(labels, fixed, width=0.32, color=[SERIES_GRAY, BLUE_DARK],
           edgecolor=PAPER, linewidth=1.5, label="Fixed opening cost")
    ax.bar(labels, transport, bottom=fixed, width=0.32,
           color=[GRAY_FILL, BLUE_SOFT], edgecolor=PAPER, linewidth=1.5,
           label="Transport cost")

    for k in range(2):
        total = fixed[k] + transport[k]
        ax.text(k, total, f"${total:,.0f}", ha="center", va="bottom",
                fontsize=11, fontweight="bold", color=INK)

    ax.set_ylabel("Cost ($)")
    ax.set_ylim(0, max(f + t for f, t in zip(fixed, transport, strict=True)) * 1.18)
    plates.money_axis(ax)
    ax.grid(True, axis="y")
    ax.set_axisbelow(True)

    # Legend swatches show the MILP blues; the caption explains the gray bar.
    handles = [
        plt.Rectangle((0, 0), 1, 1, facecolor=BLUE_DARK, edgecolor=PAPER,
                      label="Fixed opening cost"),
        plt.Rectangle((0, 0), 1, 1, facecolor=BLUE_SOFT, edgecolor=PAPER,
                      label="Transport cost"),
    ]
    ax.legend(handles=handles, loc="upper left", bbox_to_anchor=(1.08, 1.0),
              borderaxespad=0.0)

    plates.footer(fig, "The baseline bar is de-emphasized in gray; its stack "
                  "order matches the legend (fixed opening cost below, transport "
                  "above). Synthetic, seeded data; model-based estimates.")
    pdf.savefig(fig)
    plt.close(fig)


def _pooling_chart(pdf: PdfPages, r: PipelineResult) -> None:
    p = r.pooling
    fig = plates.new_plate(
        4, f"Risk pooling at {p.service_level:.0%} service (z={p.z:.2f})",
        f"-{p.network_reduction_pct:.0f}% network, "
        f"-{p.reduction_pct:.0f}% full centralization",
    )
    ax = fig.add_axes([0.13, 0.17, 0.74, 0.58])
    plates.despine(ax)
    labels = [
        "Decentralized\n(1 stock point / zone)",
        f"Network\n({r.milp.n_opened} opened DCs)",
        "Fully centralized\n(single DC)",
    ]
    values = [p.decentralized, p.network, p.centralized]
    # Entity colors held constant across the atlas: decentralized gray,
    # the opened-DC network blue, full centralization aqua.
    colors = [SERIES_GRAY, BLUE, AQUA]
    bars = ax.bar(labels, values, width=0.5, color=colors)
    for bar, v in zip(bars, values, strict=True):
        ax.text(bar.get_x() + bar.get_width() / 2, v, f"{v:,.0f}",
                ha="center", va="bottom", fontsize=11, fontweight="bold",
                color=INK)

    ax.set_ylabel("Total safety stock (units)")
    ax.set_ylim(0, max(values) * 1.15)
    plates.comma_axis(ax)
    ax.grid(True, axis="y")
    ax.set_axisbelow(True)
    ax.tick_params(axis="x", labelcolor=INK2, labelsize=9)

    plates.footer(fig, "Safety stock is modeled (normal, independent demand; "
                  "fixed lead times), so the pooling gain is an optimistic "
                  "bound. Synthetic, seeded data.")
    pdf.savefig(fig)
    plt.close(fig)


def _co2_pareto(pdf: PdfPages, r: PipelineResult) -> None:
    s = r.co2
    fig = plates.new_plate(5, "Cost vs CO2 by network density",
                           "(illustrative emission factor)")
    ax = fig.add_axes([0.09, 0.18, 0.83, 0.60])
    plates.despine(ax)

    # Pareto frontier connector (sorted by cost).
    front = sorted(s.pareto, key=lambda d: d.cost)
    if len(front) >= 2:
        ax.plot([d.cost for d in front], [d.co2_t for d in front],
                color=BLUE, linewidth=1.4, alpha=0.65, zorder=1,
                label="Pareto frontier")

    # Pareto designs wear blue; dominated designs recede to gray.
    for d in s.designs:
        color = BLUE if d.is_pareto else SERIES_GRAY
        is_opt = d is s.cost_optimal
        ax.scatter(d.cost, d.co2_t, s=150 if is_opt else 100, c=color,
                   edgecolors=INK if is_opt else PAPER,
                   linewidths=1.2 if is_opt else 1.5, zorder=3)
        ax.annotate(f"{d.n_dc} DCs", (d.cost, d.co2_t), fontsize=9,
                    color=INK2, xytext=(7, 7), textcoords="offset points")

    opt = s.cost_optimal
    ax.annotate("cost-optimal", (opt.cost, opt.co2_t), fontsize=9,
                color=INK2, xytext=(7, -15), textcoords="offset points")

    ax.set_xlabel("Total cost ($) - lower is better")
    ax.set_ylabel("Modeled outbound CO2 (tonnes) - lower is better")
    plates.money_axis(ax, axis="x")
    ax.grid(True)
    ax.set_axisbelow(True)
    ax.margins(x=0.07, y=0.08)

    handles = [
        Line2D([], [], marker="o", linestyle="none", markersize=8,
               markerfacecolor=BLUE, markeredgecolor=PAPER,
               label="Pareto-optimal"),
    ]
    if any(not d.is_pareto for d in s.designs):
        handles.append(Line2D([], [], marker="o", linestyle="none", markersize=7,
                              markerfacecolor=SERIES_GRAY, markeredgecolor=PAPER,
                              label="dominated"))
    if len(front) >= 2:
        handles.insert(0, Line2D([], [], color=BLUE, linewidth=1.4, alpha=0.65,
                                 label="Pareto frontier"))
    ax.legend(handles=handles, loc="upper right")

    note = "  ".join([
        "Point label = number of open DCs.",
        f"CO2 factor {s.emission_factor_kg_per_tkm:.2f} kg/t-km "
        f"({s.unit_weight_t:.2f} t/unit) is ILLUSTRATIVE, not certified.",
    ])
    read = tradeoff_readout(s)
    caption = [note] + ([read[1]] if len(read) > 1 else [])
    plates.footer(fig, caption)
    pdf.savefig(fig)
    plt.close(fig)


def _resilience_chart(pdf: PdfPages, r: PipelineResult) -> None:
    res = r.resilience
    verdict = ("N-1 resilient" if res.n_1_resilient
               else f"{res.n_critical} of {len(res.scenarios)} opened DCs are critical")
    fig = plates.new_plate(6, "Disruption resilience: single-DC-outage screen",
                           f"({verdict})")
    ax_fill = fig.add_axes([0.07, 0.20, 0.39, 0.54])
    ax_prem = fig.add_axes([0.56, 0.20, 0.37, 0.54])
    plates.despine(ax_fill)
    plates.despine(ax_prem)
    labels = [f"lose {s.failed_dc}" for s in res.scenarios]
    fills = [s.fill_rate_pct for s in res.scenarios]
    premiums = [s.recovery_premium for s in res.scenarios]

    # Left: each outage as served-vs-unmet demand, stacked to the 100% mark.
    # Status red is reserved for the genuinely bad part: the unmet slice.
    unmet = [100.0 - v for v in fills]
    bars = ax_fill.bar(labels, fills, width=0.5, color=BLUE, edgecolor=PAPER,
                       linewidth=1.5, label="served")
    ax_fill.bar(labels, unmet, bottom=fills, width=0.5, color=CRITICAL,
                edgecolor=PAPER, linewidth=1.5, label="unmet")
    for bar, v in zip(bars, fills, strict=True):
        ax_fill.text(bar.get_x() + bar.get_width() / 2, v - 2.5, f"{v:.1f}%",
                     ha="center", va="top", fontsize=10, fontweight="bold",
                     color=PAPER)
    ax_fill.set_ylim(0, 100)
    ax_fill.set_ylabel("Demand served within surviving fleet (%)")
    ax_fill.set_title("N-1 outage: fill rate if one DC is lost", pad=10)
    ax_fill.grid(True, axis="y")
    ax_fill.set_axisbelow(True)
    ax_fill.legend(loc="upper center", bbox_to_anchor=(0.5, -0.10), ncols=2)

    # Right: cheapest recovery premium to restore 100% service.
    pbars = ax_prem.bar(labels, premiums, width=0.5, color=BLUE)
    for bar, s in zip(pbars, res.scenarios, strict=True):
        tag = ("+" + ", ".join(s.recovery_activated)) if s.recovery_activated else "no add"
        ax_prem.text(bar.get_x() + bar.get_width() / 2, s.recovery_premium,
                     f"${s.recovery_premium:,.0f}\n{tag}", ha="center",
                     va="bottom", fontsize=9, color=INK)
    ax_prem.set_ylabel("Recovery premium ($ added fixed + transport)")
    ax_prem.set_title("Cheapest standby activation to restore 100%", pad=10)
    plates.money_axis(ax_prem)
    ax_prem.grid(True, axis="y")
    ax_prem.set_axisbelow(True)
    ax_prem.margins(y=0.22)

    plates.footer(fig, resilience_readout(res)[1])
    pdf.savefig(fig)
    plt.close(fig)


def _service_frontier_chart(pdf: PdfPages, r: PipelineResult) -> None:
    sf = r.service_frontier
    fig = plates.new_plate(7, "Inventory service-level frontier",
                           "(illustrative inventory-cost factors)")
    ax_ss = fig.add_axes([0.07, 0.18, 0.39, 0.56])
    ax_marg = fig.add_axes([0.57, 0.18, 0.36, 0.56])
    plates.despine(ax_ss)
    plates.despine(ax_marg)
    svc = [p.service_level * 100.0 for p in sf.points]

    # Left: safety stock vs service level under the three pooling regimes.
    # The regimes keep their atlas colors; the network line carries the weight.
    ax_ss.plot(svc, [p.ss_decentralized for p in sf.points], marker="o",
               markersize=4.5, markeredgecolor=PAPER, markeredgewidth=1.0,
               color=SERIES_GRAY, linewidth=1.8, linestyle=(0, (4, 2)),
               label="Decentralized (1 point / zone)")
    ax_ss.plot(svc, [p.ss_centralized for p in sf.points], marker="o",
               markersize=4.5, markeredgecolor=PAPER, markeredgewidth=1.0,
               color=AQUA, linewidth=1.8, label="Fully centralized")
    ax_ss.plot(svc, [p.ss_network for p in sf.points], marker="o",
               markersize=5.5, markeredgecolor=PAPER, markeredgewidth=1.2,
               color=BLUE, linewidth=2.4,
               label=f"Network ({r.milp.n_opened} opened DCs)")
    ax_ss.set_xlabel("Target service level (%)")
    ax_ss.set_ylabel("Safety stock (units)")
    plates.comma_axis(ax_ss)
    ax_ss.set_title("Safety stock rises convexly with the service target",
                    pad=10)
    ax_ss.legend(loc="upper left")
    ax_ss.grid(True)
    ax_ss.set_axisbelow(True)

    # Right: marginal carrying cost of each extra service point (network regime).
    marg_pts = [p for p in sf.points if p.marginal_cost_per_point > 0]
    labels = [f"{p.service_level:.1%}" for p in marg_pts]
    marginals = [p.marginal_cost_per_point for p in marg_pts]
    bars = ax_marg.bar(labels, marginals, width=0.55, color=BLUE)
    for bar, v in zip(bars, marginals, strict=True):
        ax_marg.text(bar.get_x() + bar.get_width() / 2, v, f"${v:,.0f}",
                     ha="center", va="bottom", fontsize=9, color=INK)
    ax_marg.set_xlabel("Service level reached")
    ax_marg.set_ylabel("Marginal carrying cost ($/yr per service point)")
    ax_marg.set_title("The last points of service cost the most", pad=10)
    plates.money_axis(ax_marg)
    ax_marg.grid(True, axis="y")
    ax_marg.set_axisbelow(True)
    ax_marg.margins(y=0.18)

    plates.footer(fig, frontier_readout(sf)[2])
    pdf.savefig(fig)
    plt.close(fig)


def _growth_chart(pdf: PdfPages, r: PipelineResult) -> None:
    g = r.growth
    fig = plates.new_plate(
        8, "Demand-growth capacity plan: expansion triggers and headroom",
        "(uniform growth, synthetic data)",
    )
    # The staircase is the hero panel, on top; the cost sweep sits beneath it
    # on the same growth axis. Two stacked panels -- never a dual axis.
    ax_stair = fig.add_axes([0.09, 0.52, 0.84, 0.28])
    ax_cost = fig.add_axes([0.09, 0.16, 0.84, 0.26], sharex=ax_stair)
    plates.despine(ax_stair)
    plates.despine(ax_cost)
    growths = [p.growth for p in g.points]

    # Top: the expansion staircase -- optimal open-DC count vs demand growth.
    ax_stair.step(growths, [p.n_opened for p in g.points], where="post",
                  color=BLUE, linewidth=2.6, zorder=3)
    for p in g.expansion_triggers:
        ax_stair.scatter([p.growth], [p.n_opened], s=42, c=BLUE,
                         edgecolors=PAPER, linewidths=1.4, zorder=4)
        ax_stair.annotate(f"DC #{p.n_opened} pays\nat {p.growth:.2f}x",
                          (p.growth, p.n_opened), fontsize=8, color=INK2,
                          xytext=(6, -24), textcoords="offset points")
    ax_stair.set_ylabel("Optimal number of open DCs")
    ax_stair.set_yticks(sorted({p.n_opened for p in g.points}))
    ax_stair.set_title("The expansion staircase (MILP re-solved per level)",
                       pad=10)
    ax_stair.grid(True)
    ax_stair.set_axisbelow(True)
    ax_stair.margins(y=0.25)
    ax_stair.tick_params(labelbottom=False)

    # Bottom: re-optimized vs frozen committed cost across the growth sweep.
    ax_cost.plot(growths, [p.cost for p in g.points], marker="o", markersize=4.5,
                 markeredgecolor=PAPER, markeredgewidth=1.0, color=BLUE,
                 linewidth=2.2, label="Re-optimized at each level", zorder=3)
    committed = [p for p in g.points if p.committed_feasible]
    if committed:
        # Up to the wall the frozen cost coincides with the re-optimized cost,
        # so the frozen series rides ON the blue line: hollow gray rings on
        # top keep it visible exactly where the two series overlap.
        ax_cost.plot([p.growth for p in committed],
                     [p.committed_cost for p in committed], marker="o",
                     markersize=7.5, markerfacecolor="none",
                     markeredgecolor=SERIES_GRAY, markeredgewidth=1.6,
                     color=SERIES_GRAY, linewidth=1.8, linestyle=(0, (4, 2)),
                     label="Committed network frozen", zorder=4)
    ax_cost.set_xlabel("Demand growth (x today)")
    ax_cost.set_ylabel("Total network cost ($, fixed + outbound transport)",
                       fontsize=8)
    ax_cost.set_title("Cost of growth: redesign vs frozen network", pad=10)
    plates.money_axis(ax_cost)
    ax_cost.legend(loc="lower right")
    ax_cost.grid(True)
    ax_cost.set_axisbelow(True)

    # The committed capacity wall, marked on both panels; labeled once, on the
    # hero panel. Status red: past this line the frozen network is infeasible.
    for ax in (ax_stair, ax_cost):
        ax.axvline(g.committed_wall_growth, color=CRITICAL,
                   linestyle=(0, (4, 2)), linewidth=1.2, zorder=2)
    ax_stair.annotate(
        f"committed wall {g.committed_wall_growth:.3f}x\n"
        f"({g.committed_headroom_pct:+.1f}% headroom)",
        (g.committed_wall_growth, ax_stair.get_ylim()[1]), fontsize=8,
        color=CRITICAL, ha="left", va="top",
        xytext=(6, -2), textcoords="offset points",
    )

    plates.footer(fig, growth_readout(g)[0])
    pdf.savefig(fig)
    plt.close(fig)


def build_pdf(r: PipelineResult, path: str) -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    plates.apply_style()
    with PdfPages(path) as pdf:
        _cover_page(pdf, r)
        _network_map(pdf, r)
        _cost_breakdown(pdf, r)
        _pooling_chart(pdf, r)
        _co2_pareto(pdf, r)
        _resilience_chart(pdf, r)
        _service_frontier_chart(pdf, r)
        _growth_chart(pdf, r)
        meta = pdf.infodict()
        meta["Title"] = "Supply-Network Optimization (synthetic)"
        meta["Author"] = "Dimitres Kisimov"
    return path


def build_excel(r: PipelineResult, path: str) -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Supply-Network Optimization - Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "SYNTHETIC data, seed " + str(r.data.seed) + " - model-based estimates"
    rows = [
        ("Metric", "Value"),
        ("Plants / candidate DCs / customers",
         f"{r.data.n_plants} / {r.data.n_dcs} / {r.data.n_customers}"),
        ("Total demand (units)", round(r.data.total_demand, 1)),
        ("DCs opened (MILP)", r.milp.n_opened),
        ("Opened DC ids", ", ".join(r.milp.opened)),
        ("MILP total cost ($)", round(r.milp.total_cost, 2)),
        ("MILP fixed cost ($)", round(r.milp.fixed_cost, 2)),
        ("MILP transport cost ($)", round(r.milp.transport_cost, 2)),
        ("Greedy baseline cost ($)", round(r.greedy.total_cost, 2)),
        ("Savings vs baseline ($)", round(r.savings_abs, 2)),
        ("Savings vs baseline (%)", round(r.savings_pct, 2)),
        ("Min-cost flow (graph) ($)", round(r.flow["graph"].total_cost, 2)),
        ("Min-cost flow (LP) ($)", round(r.flow["lp"].total_cost, 2)),
        ("Flow graph-vs-LP abs gap ($)", round(r.flow["abs_gap"], 6)),
        ("Service level", r.service_level),
        ("Safety-stock z", round(r.pooling.z, 4)),
        ("Safety stock decentralized (units)", round(r.pooling.decentralized, 1)),
        ("Safety stock network (units)", round(r.pooling.network, 1)),
        ("Safety stock centralized (units)", round(r.pooling.centralized, 1)),
        ("Pooling reduction network (%)", round(r.pooling.network_reduction_pct, 2)),
        ("Pooling reduction full (%)", round(r.pooling.reduction_pct, 2)),
        ("N-1 resilient (any single DC loss)",
         "yes" if r.resilience.n_1_resilient else "no"),
        ("Critical DCs (loss drops service)", r.resilience.n_critical),
        ("Worst single-loss fill rate (%)",
         round(r.resilience.worst.fill_rate_pct, 2)),
        ("Max recovery premium ($)", round(r.resilience.max_recovery_premium, 2)),
        ("Safety-stock carrying cost @ base SL ($/yr)",
         round(r.service_frontier.base_point.holding_cost_network, 2)),
        ("Marginal cost of last service point ($/yr per pt)",
         round(r.service_frontier.steepest.marginal_cost_per_point, 2)),
        ("Last-point vs first-point cost ratio (x)",
         round(r.service_frontier.marginal_ratio_top_to_bottom, 2)),
        ("Committed growth headroom (%)",
         round(r.growth.committed_headroom_pct, 2)),
        ("Committed capacity wall (x demand)",
         round(r.growth.committed_wall_growth, 4)),
        ("First expansion trigger (x demand)",
         r.growth.first_expansion.growth if r.growth.first_expansion else "n/a"),
        ("DCs opened at max swept growth",
         r.growth.points[-1].n_opened if r.growth.points else "n/a"),
        ("DC-portfolio growth ceiling (x demand)",
         round(r.growth.dc_ceiling_growth, 4)),
    ]
    for ridx, (a, b) in enumerate(rows, start=4):
        ws.cell(ridx, 1, a)
        ws.cell(ridx, 2, b)
        if ridx == 4:
            ws.cell(ridx, 1).font = bold
            ws.cell(ridx, 2).font = bold
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30

    # Facilities sheet.
    wf = wb.create_sheet("Facilities")
    wf.append(["dc_id", "x", "y", "capacity", "fixed_cost", "opened",
               "throughput_used"])
    for c in wf[1]:
        c.font = bold
    used = r.milp.assignment.sum(axis=1)
    for i, row in r.data.dcs.iterrows():
        wf.append([
            row["dc_id"], round(float(row["x"]), 2), round(float(row["y"]), 2),
            float(row["capacity"]), float(row["fixed_cost"]),
            "yes" if row["dc_id"] in set(r.milp.opened) else "no",
            round(float(used[i]), 1),
        ])
    for col in "ABCDEFG":
        wf.column_dimensions[col].width = 14

    # Flows sheet (multi-echelon arc flows from the graph solver).
    wflow = wb.create_sheet("Flows")
    wflow.append(["arc", "flow_units"])
    for c in wflow[1]:
        c.font = bold
    for label, val in r.flow["graph"].arc_flows:
        wflow.append([label, round(val, 1)])
    wflow.column_dimensions["A"].width = 34
    wflow.column_dimensions["B"].width = 14

    # SafetyStock sheet.
    wss = wb.create_sheet("SafetyStock")
    wss.append(["scenario", "safety_stock_units", "note"])
    for c in wss[1]:
        c.font = bold
    p = r.pooling
    wss.append(["decentralized", round(p.decentralized, 1),
                "one stock point per customer zone"])
    wss.append(["network", round(p.network, 1),
                f"pooled by the {r.milp.n_opened} opened DCs"])
    wss.append(["centralized", round(p.centralized, 1), "single pooled DC"])
    wss.append([])
    wss.append(["echelon", "safety_stock_units", "lead_time_days"])
    wss.append(["DC echelon", round(r.echelon["dc_echelon"], 1),
                round(r.echelon["avg_lead_time"], 2)])
    wss.append(["Plant echelon", round(r.echelon["plant_echelon"], 1),
                round(r.echelon["avg_lead_time"] + 7.0, 2)])
    wss.column_dimensions["A"].width = 20
    wss.column_dimensions["B"].width = 20
    wss.column_dimensions["C"].width = 34

    # Customers sheet: the full synthetic demand table.
    wc = wb.create_sheet("Customers")
    wc.append(["cust_id", "x", "y", "demand_mean", "demand_std", "lead_time",
               "served_by"])
    for c in wc[1]:
        c.font = bold
    served_by = r.milp.assignment.argmax(axis=0)
    dc_ids = list(r.data.dcs["dc_id"])
    for j, row in r.data.customers.iterrows():
        wc.append([
            row["cust_id"], round(float(row["x"]), 2), round(float(row["y"]), 2),
            float(row["demand_mean"]), float(row["demand_std"]),
            float(row["lead_time"]), dc_ids[int(served_by[j])],
        ])
    for col in "ABCDEFG":
        wc.column_dimensions[col].width = 13

    # CO2 sensitivity sheet: cost-optimal design at each network density.
    wco2 = wb.create_sheet("CO2Sensitivity")
    s = r.co2
    wco2["A1"] = "CO2 / cost / service sensitivity (sweep over # open DCs)"
    wco2["A1"].font = Font(bold=True, size=12)
    wco2["A2"] = (
        "Emission factor ILLUSTRATIVE: "
        f"{s.emission_factor_kg_per_tkm:.2f} kg CO2e/tonne-km, "
        f"{s.unit_weight_t:.2f} t/unit - NOT a certified figure. Outbound leg only."
    )
    hdr = ["n_dcs", "opened", "cost_usd", "co2_tonnes", "avg_delivery_km",
           f"pct_demand_within_{int(s.service_radius_km)}km", "pareto_optimal"]
    wco2.append([])
    wco2.append(hdr)
    for c in wco2[4]:
        c.font = bold
    for d in s.designs:
        wco2.append([
            d.n_dc, ", ".join(d.opened), round(d.cost, 2), round(d.co2_t, 4),
            round(d.avg_delivery_km, 2), round(d.service_within_radius_pct, 2),
            "yes" if d.is_pareto else "no",
        ])
    wco2.append([])
    for line in tradeoff_readout(s):
        wco2.append([line])
    wco2.column_dimensions["A"].width = 16
    for col in "BCDEFG":
        wco2.column_dimensions[col].width = 16

    # Resilience sheet: single-DC-outage (N-1) screen + cheapest recovery.
    wres = wb.create_sheet("Resilience")
    res = r.resilience
    wres["A1"] = "Disruption resilience - single-DC-outage (N-1) screen + recovery"
    wres["A1"].font = Font(bold=True, size=12)
    verdict = ("N-1 resilient (survives any single DC loss at 100%)"
               if res.n_1_resilient
               else f"NOT N-1 resilient: {res.n_critical} of "
                    f"{len(res.scenarios)} opened DCs are critical")
    wres["A2"] = f"Committed network {', '.join(res.committed_opened)} - {verdict}"
    hdr = ["lost_dc", "surviving_dcs", "surviving_capacity", "fill_rate_pct",
           "unmet_units", "reroute_transport_usd", "recovery_activate",
           "recovery_added_fixed_usd", "recovery_premium_usd", "recovery_possible"]
    wres.append([])
    wres.append(hdr)
    for c in wres[4]:
        c.font = bold
    for s in res.scenarios:
        wres.append([
            s.failed_dc, ", ".join(s.survivors), round(s.surviving_capacity, 1),
            round(s.fill_rate_pct, 2), round(s.unmet_units, 1),
            round(s.reroute_transport_cost, 2),
            ", ".join(s.recovery_activated) if s.recovery_activated else "(none)",
            round(s.recovery_added_fixed, 2), round(s.recovery_premium, 2),
            "yes" if s.recovery_possible else "no",
        ])
    wres.append([])
    for line in resilience_readout(res):
        wres.append([line])
    wres.column_dimensions["A"].width = 16
    for col in "BCDEFGHIJ":
        wres.column_dimensions[col].width = 18

    # ServiceFrontier sheet: safety stock / inventory cost vs target service level.
    wsf = wb.create_sheet("ServiceFrontier")
    sf = r.service_frontier
    wsf["A1"] = "Inventory service-level frontier (sweep over target service level)"
    wsf["A1"].font = Font(bold=True, size=12)
    wsf["A2"] = (
        "Inventory $ ILLUSTRATIVE: "
        f"${sf.unit_value_usd:,.0f}/unit value, {sf.holding_rate_per_year:.0%}/yr "
        "carrying rate - NOT certified. Pooling assumes independent demand."
    )
    hdr = ["service_level", "z", "ss_decentralized", "ss_network",
           "ss_centralized", "inv_value_network_usd", "holding_cost_usd_per_yr",
           "marginal_usd_per_service_point"]
    wsf.append([])
    wsf.append(hdr)
    for c in wsf[4]:
        c.font = bold
    for p in sf.points:
        wsf.append([
            round(p.service_level, 4), round(p.z, 4),
            round(p.ss_decentralized, 1), round(p.ss_network, 1),
            round(p.ss_centralized, 1), round(p.inv_value_network, 2),
            round(p.holding_cost_network, 2), round(p.marginal_cost_per_point, 2),
        ])
    wsf.append([])
    for line in frontier_readout(sf):
        wsf.append([line])
    wsf.column_dimensions["A"].width = 16
    for col in "BCDEFGH":
        wsf.column_dimensions[col].width = 18

    # Growth sheet: demand-growth expansion plan (re-optimized vs frozen).
    wg = wb.create_sheet("Growth")
    g = r.growth
    wg["A1"] = "Demand-growth capacity plan (sweep over a uniform growth multiplier)"
    wg["A1"].font = Font(bold=True, size=12)
    wg["A2"] = (
        "UNIFORM growth (every zone scales alike), deterministic demand, flat "
        "costs and rates; DC capacity is the only hard limit. Committed columns "
        "blank past the capacity wall."
    )
    hdr = ["growth_x", "demand_units", "n_dcs_optimal", "opened", "cost_usd",
           "cost_per_unit_usd", "capacity_utilization_pct", "reconfigured",
           "expanded", "committed_feasible", "committed_cost_usd",
           "committed_premium_usd"]
    wg.append([])
    wg.append(hdr)
    for c in wg[4]:
        c.font = bold
    for p in g.points:
        wg.append([
            round(p.growth, 2), round(p.demand_units, 1), p.n_opened,
            ", ".join(p.opened), round(p.cost, 2), round(p.cost_per_unit, 4),
            round(p.utilization_pct, 2), "yes" if p.reconfigured else "no",
            "yes" if p.expanded else "no",
            "yes" if p.committed_feasible else "no",
            round(p.committed_cost, 2) if p.committed_feasible else "",
            round(p.committed_premium, 2) if p.committed_feasible else "",
        ])
    wg.append([])
    for line in growth_readout(g):
        wg.append([line])
    wg.column_dimensions["A"].width = 12
    for col in "BCDEFGHIJKL":
        wg.column_dimensions[col].width = 18

    # Assignment sheet: opened-DC x customer shipped-units matrix.
    wa = wb.create_sheet("Assignment")
    header = ["dc_id \\ cust", *list(r.data.customers["cust_id"])]
    wa.append(header)
    for c in wa[1]:
        c.font = bold
    for i, dc_id in enumerate(dc_ids):
        if dc_id not in set(r.milp.opened):
            continue
        wa.append([dc_id, *[round(float(v), 1) for v in r.milp.assignment[i]]])
    wa.column_dimensions["A"].width = 14

    wb.save(path)
    return path


def write_deliverables(r: PipelineResult | None = None, out_dir: str = "deliverables") -> dict:
    """Write both deliverables to ``out_dir`` and return their paths."""
    if r is None:
        r = run_pipeline()
    os.makedirs(out_dir, exist_ok=True)
    pdf_path = os.path.join(out_dir, "supply_network_report.pdf")
    xlsx_path = os.path.join(out_dir, "supply_network_workbook.xlsx")
    csv_path = os.path.join(out_dir, "co2_sensitivity.csv")
    svg_path = os.path.join(out_dir, "co2_cost_frontier.svg")
    sf_csv_path = os.path.join(out_dir, "service_frontier.csv")
    sf_svg_path = os.path.join(out_dir, "service_frontier.svg")
    growth_csv_path = os.path.join(out_dir, "growth_plan.csv")
    growth_svg_path = os.path.join(out_dir, "growth_expansion.svg")
    build_pdf(r, pdf_path)
    build_excel(r, xlsx_path)
    with open(csv_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(to_csv(r.co2))
    with open(svg_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(to_svg(r.co2))
    with open(sf_csv_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(frontier_to_csv(r.service_frontier))
    with open(sf_svg_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(frontier_to_svg(r.service_frontier))
    with open(growth_csv_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(growth_to_csv(r.growth))
    with open(growth_svg_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(growth_to_svg(r.growth))
    return {
        "pdf": pdf_path,
        "xlsx": xlsx_path,
        "csv": csv_path,
        "svg": svg_path,
        "sf_csv": sf_csv_path,
        "sf_svg": sf_svg_path,
        "growth_csv": growth_csv_path,
        "growth_svg": growth_svg_path,
    }
